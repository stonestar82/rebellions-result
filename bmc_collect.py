from openpyxl import load_workbook
import os
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
import time
from PIL import Image
from datetime import datetime
import io
import os
    
class ResultProcess:
	def __init__(self):
		pass


	def capture_element_screenshot(self, driver, ip, xpath, element_name):
		"""XPath로 지정된 요소의 스크린샷을 캡처하는 함수"""
		try:
			## 요소가 존재하는지 확인
			element = driver.find_element(By.XPATH, xpath)
			
			## 요소의 위치와 크기 가져오기
			location = element.location
			size = element.size
	
			## 전체 스크린샷
			temp_filename = f"{self.folder_name}/screenshot_{element_name}.png"
			driver.save_screenshot(temp_filename)
	
			## 스크린샷을 PIL Image로 변환
			screenshot = Image.open(temp_filename)
			
			## 요소 영역만 크롭
			left = location['x']
			top = location['y']
			right = location['x'] + size['width']
			bottom = location['y'] + size['height']
			
			element_screenshot = screenshot.crop((left, top, right, bottom))
			## 크롭된 이미지 저장
			final_filename = f"{self.folder_name}/{element_name}.png"
			element_screenshot.save(final_filename)
			
			print(f"요소 스크린샷 저장 완료: {final_filename}")
			
			## 전체 스크린샷 파일 삭제
			os.remove(temp_filename)
			
		except Exception as e:
			raise Exception("스크린샷 캡처 실패 : " + e)
   
  
	def get_result(self):
		# print("get_result")
			
		# Chrome 옵션 설정 - SSL 인증서 오류 무시
		chrome_options = Options()
		chrome_options.add_argument("--no-sandbox")
		chrome_options.add_argument("--disable-dev-shm-usage")
		chrome_options.add_argument("--disable-gpu")
		chrome_options.add_argument("--ignore-ssl-errors")
		chrome_options.add_argument("--ignore-certificate-errors")
		chrome_options.add_argument("--ignore-certificate-errors-spki-list")
		chrome_options.add_argument("--ignore-ssl-errors-spki-list")
		chrome_options.add_argument("--allow-insecure-localhost")
		chrome_options.add_argument("--disable-web-security")
		chrome_options.add_argument("--allow-running-insecure-content")
		chrome_options.add_argument("--disable-features=VizDisplayCompositor")
		
		# 추가: 사용자 프로필 설정으로 인증서 오류 무시
		chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
		chrome_options.add_experimental_option('useAutomationExtension', False)

		chrome_path = ChromeDriverManager().install()
		if "THIRD_PARTY_NOTICES.chromedriver" in chrome_path:
			print("THIRD_PARTY_NOTICES.chromedriver")
			chrome_path = chrome_path.replace("THIRD_PARTY_NOTICES.chromedriver", "chromedriver.exe")
  
		# chrome_options.binary_location = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
		service = Service(chrome_path)
		driver = webdriver.Chrome(service=service, options=chrome_options)
  
		wait = WebDriverWait(driver, 10)
  
	

		wb = load_workbook(filename="inventory.xlsx")

		sheet = wb["bmc"]

		# 두번째 행부터 마지막 행까지 출력
		for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=True):
			ip = row[0]
			id = row[1]
			pw = row[2]
			name = row[3]
			check = str(row[4]).upper()
			manufacturer = row[5]
			product_name = row[6]
			serial_number = row[7]
   
			if check == "N":
				print(f"skip {ip}")
				time.sleep(2)
				continue
			
   
			## 오늘날짜 + name 폴더
			self.folder_name = f"{manufacturer}_{product_name}_{serial_number}"
			os.makedirs(self.folder_name, exist_ok=True)
   
			## 오늘날짜 + name 폴더
			self.folder_name = f"{self.folder_name}/bmc"
			os.makedirs(self.folder_name, exist_ok=True)
   
			print("========================")
			print("bmc ip : ", ip)
			print("collect start")
			print("========================")
   
			driver.get("https://" + ip)
   
			## 전체화면 으로 변경
			driver.maximize_window()
   
			driver.find_element(By.ID, "usrName").send_keys(id)
			driver.find_element(By.ID, "pwd").send_keys(pw)
			driver.find_element(By.ID, "login_word").click()
   
   
   
			## sysBtn 클릭 가능할때까지 대기
			sysBtn = wait.until(EC.element_to_be_clickable((By.ID, "sysBtn")))
			## 페이지 이동후 로드 완료되기 까지 대기
			# wait.until(EC.presence_of_element_located((By.ID, "login_word")))   
   
			time.sleep(5)
   
   
			## 시스템 정보 추출
			self.capture_element_screenshot(driver, ip, "/html/body/div[1]/main/div/div/div[1]/div/div/div[2]/div[1]/div", "system")
  
			sysBtn.click()
			time.sleep(5)
	
			## cpuTab 
			cpuTab = wait.until(EC.element_to_be_clickable((By.ID, "cpuTab")))
	
			## cpuTab 클릭
			cpuTab.click()
			time.sleep(5)
   
			## cpu 정보 추출
			self.capture_element_screenshot(driver, ip, "/html/body/div[1]/main/div/div/div[1]/div/div[2]", "cpu")
	
			
			## memory 정보 추출
			memoryTab = wait.until(EC.element_to_be_clickable((By.ID, "memoryTab")))
			memoryTab.click()
			time.sleep(5)

			## memory 정보 추출
			self.capture_element_screenshot(driver, ip, "/html/body/div[1]/main/div/div/div[1]/div/div[3]", "memory")


			## psu 정보 추출
			psuTab = wait.until(EC.element_to_be_clickable((By.ID, "psuTab")))
			psuTab.click()
			time.sleep(5)
   
			## psu 정보 추출
			self.capture_element_screenshot(driver, ip, "/html/body/div[1]/main/div/div/div[1]/div/div[4]", "psu")
   
			## nic 정보 추출
			nicTab = wait.until(EC.element_to_be_clickable((By.ID, "networkTab")))
			nicTab.click()
			time.sleep(5)
   
			tbl = driver.find_element(By.XPATH, "/html/body/div[1]/main/div/div/div[1]/div/div[7]/div/div[2]/div[2]")
   
			a_tag = tbl.find_elements(By.TAG_NAME, "a")
			for a in a_tag:
				a.click()
				time.sleep(2)
   
			## nic 정보 추출
			self.capture_element_screenshot(driver, ip, "/html/body/div[1]/main/div/div/div[1]/div/div[7]", "nic")
   
   
			## npu 정보 추출
			gpuTab = wait.until(EC.element_to_be_clickable((By.ID, "gpuTab")))
			gpuTab.click()
			time.sleep(5)

			## npu 정보 추출
			self.capture_element_screenshot(driver, ip, "/html/body/div[1]/main/div/div/div[1]/div/div[10]", "npu")
   
   
			## fan 정보 추출
			coolingTab = wait.until(EC.element_to_be_clickable((By.ID, "coolingTab")))
			coolingTab.click()
			time.sleep(5)
   
			## fan 정보 추출
			self.capture_element_screenshot(driver, ip, "/html/body/div[1]/main/div/div/div[1]/div/div[9]", "fan_speed")

	

			## 메인 이동후 RAID CONTROLLER 정보 추출
			main_tab = driver.find_element(By.XPATH, "/html/body/header/a")
			main_tab.click()
			time.sleep(5)
	
			storage_tab = wait.until(EC.element_to_be_clickable((By.ID, "storageBtn")))
			storage_tab.click()
			time.sleep(5)
   
			ctrl_tab = wait.until(EC.element_to_be_clickable((By.ID, "Controllers_tab")))
			ctrl_tab.click()
			time.sleep(5)
   
			expanstion_tab = driver.find_element(By.XPATH, "/html/body/div[1]/main/div/div/div[1]/div[2]/div[4]/div[3]/div/div/div[1]/div/span[2]")
			expanstion_tab.click()
			time.sleep(2)

			## raid 정보 추출
			self.capture_element_screenshot(driver, ip, "/html/body/div[1]/main/div/div/div[1]/div[2]/div[4]/div[3]", "raid")



    
		
process = ResultProcess()

process.get_result()



    