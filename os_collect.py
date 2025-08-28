from openpyxl import load_workbook
import os
import paramiko
from paramiko import SSHClient, AutoAddPolicy

import time
import re
from datetime import datetime
import pyautogui

class SSHExpect:
	def __init__(self, shell):
		self.shell = shell
		self.buffer = ""
	
	def expect(self, pattern, timeout=10):
		"""패턴을 기다리고 매칭된 텍스트 반환"""
		start_time = time.time()
		
		while time.time() - start_time < timeout:
				if self.shell.recv_ready():
						chunk = self.shell.recv(1024).decode()
						self.buffer += chunk
						print(chunk, end='')
						
						match = re.search(pattern, self.buffer)
						if match:
								return match.group(0)
				
				time.sleep(0.1)
		
		return None

	def send(self, text):
		"""텍스트 전송"""
		self.shell.send(text + '\n')
		return self.shell.recv(4096).decode()

	def execute_and_wait(self, command, timeout=10):
		"""명령어 실행 후 결과 대기"""
		# 명령어 전송
		self.shell.send(command + '\n')
		
		# 결과 수집
		result = ""
		start_time = time.time()
		
		prompt = ""
		while time.time() - start_time < timeout:
				if self.shell.recv_ready():
						chunk = self.shell.recv(1024).decode()
     
						result += chunk
      
						# 프롬프트가 나타났는지 확인
						m = re.search(r'root@[\w\-]+:.*[#\$]', chunk)
						if m:
								prompt = m.group(0)
								break
      
				time.sleep(0.1)
    
		result = re.sub(r'root@[\w\-]+:.*[#\$]\s*', "", result)  # root@hostname:~# 
		result = re.sub(r'[\w\-]+@[\w\-]+:.*[#\$]\s*', "", result)  # user@hostname:~$
		result = re.sub(r'[\w\-]+:.*[#\$]\s*', "", result)  # hostname:~#
	
  
		result = f"{prompt} {result}"
  
		return result

def cmd_and_capture(expect, command, folder_name, filename):
	"""
	SSHExpect.send를 사용하여 명령어 실행 후 결과 출력 및 스크린샷 촬영
	
	Args:
			expect: SSHExpect 객체
			command: 실행할 명령어
	"""
	os.system("cls")
	
	# SSHExpect.send를 사용하여 명령어 실행
	result = expect.execute_and_wait(command)
	
	# 결과 출력
	if result:
			# print("결과 출력")
			print(result)
	else:
			print("명령어 실행 결과가 없습니다.")
	
	time.sleep(3)
	
	try:
			# 현재 시간으로 타임스탬프 생성
			
			# 파일명 생성 - 명령어에서 특수문자 제거
			safe_command = "".join(c for c in command if c.isalnum() or c in (' ', '-', '_')).rstrip()
			filename = f"{folder_name}/{filename}.png"
			
			# 스크린샷 촬영
			screenshot = pyautogui.screenshot()
			screenshot.save(filename)
			
			print(f"스크린샷 저장 완료: {filename}")
			return filename
			
	except Exception as e:
			print(f"스크린샷 촬영 실패: {e}")
			return None
	

def ssh_connect(host, username, password, port):
	connection = SSHClient()
	connection.set_missing_host_key_policy(AutoAddPolicy())
	connection.connect(host, username=username, password=password, port=port, look_for_keys=False, allow_agent=False)
	return connection

print("=== SSH 연결 도구 ===")


wb = load_workbook(filename="inventory.xlsx")

sheet = wb["os"]

d = {}

# 두번째 행부터 마지막 행까지 출력
for row_num in range(2, sheet.max_row + 1):
	row = sheet[row_num]
	hostname = row[0].value
	username = row[1].value
	pw = row[2].value
	name = row[3].value
	check = str(row[4].value).upper()
	port = "22"
  
	if check == "N":
		print(f"skip {hostname}")
		time.sleep(2)
		continue

	## ssh 접속
	connection = ssh_connect(hostname, username, pw, port)



	os.system("cls")
	print("5초후 시작합니다. cmd 창을 최대화 해주세요.")
 



	os.system("cls")


	# 대화형 셸 시작
	shell = connection.invoke_shell()
	time.sleep(1)

	# 초기 출력 제거
	shell.recv(4096)

	# expect 객체 생성
	expect = SSHExpect(shell)

	# sudo su 명령어 실행
	expect.send("sudo su")

	# sudo 패스워드 프롬프트 대기
	sudo_prompt = expect.expect(r"\[sudo\] password for \w+:")
	if sudo_prompt:
		print("sudo 패스워드 프롬프트 감지됨!")
		expect.send(pw)

		root_prompt = expect.expect(r"root@")
		if root_prompt:
			print("root 권한 획득 성공!")
		else:
			raise Exception("root 권한 획득 실패")

	else:
		raise Exception("sudo 패스워드 프롬프트 감지 실패")

	time.sleep(3)

	os.system("cls")
 
	result = str(expect.execute_and_wait("dmidecode -t 1 | egrep 'Manufacturer|Product Name|Serial Number'"))
	# print("=== 원본 결과 ===")
	# print(repr(result))  # repr()을 사용하여 숨겨진 문자 확인
	# print("=== 결과 길이 ===")
	# print(f"결과 길이: {len(result)}")
	# print("=== 결과 내용 ===")
	# print(result)
	
	# ANSI 제어 문자와 터미널 제어 문자 제거
	import re
	clean_result = re.sub(r'\x1b\[[0-9;]*[a-zA-Z]', '', result)  # ANSI 제어 문자 제거
	clean_result = re.sub(r'\x1b\?2004[hl]', '', clean_result)   # 터미널 제어 문자 제거
	clean_result = re.sub(r'\r\n', '\n', clean_result)            # \r\n을 \n으로 변환
	clean_result = re.sub(r'\t', ' ', clean_result)               # 탭을 공백으로 변환
	
	# print("=== 정리된 결과 ===")
	# print(repr(clean_result))
	# print("=== 정리된 결과 내용 ===")
	# print(clean_result)
	
	# 정규식 매치 결과 확인 및 안전한 추출
	manufacturer_match = re.search(r"Manufacturer:\s*(.*)", clean_result)
	product_name_match = re.search(r"Product Name:\s*(.*)", clean_result)
	serial_number_match = re.search(r"Serial Number:\s*(.*)", clean_result)
	
	# print("=== 정규식 매치 결과 ===")
	# print(f"Manufacturer match: {manufacturer_match}")
	# print(f"Product Name match: {product_name_match}")
	# print(f"Serial Number match: {serial_number_match}")
	
	if manufacturer_match and product_name_match and serial_number_match:
		manufacturer = manufacturer_match.group(1).strip()
		product_name = product_name_match.group(1).strip().replace(" ", "")
		serial_number = serial_number_match.group(1).strip()
		
		print(f"Manufacturer: {manufacturer}")
		print(f"Product Name: {product_name}")
		print(f"Serial Number: {serial_number}")
  
		d.setdefault(name, {
			"manufacturer": manufacturer,
			"product_name": product_name,
			"serial_number": serial_number
		})
   
		row[5].value = manufacturer
		row[6].value = product_name
		row[7].value = serial_number
  
	else:
		print("dmidecode 결과에서 필요한 정보를 찾을 수 없습니다.")
		print("원본 결과:", result)
		print("정리된 결과:", clean_result)
		exit()
	
	

	## 오늘날짜 + name 폴더
	folder_name = f"{manufacturer}_{product_name}_{serial_number}"

	os.makedirs(folder_name, exist_ok=True)
 
	folder_name = f"{folder_name}/os"
	os.makedirs(folder_name, exist_ok=True)

	## cpu
	cmd_and_capture(expect, "lscpu | egrep 'Model name|Socket|^CPU\(s\)'", folder_name, "cpu")

	## memory
	cmd_and_capture(expect, "free -h", folder_name, "memory")

	## disk
	cmd_and_capture(expect, "lsblk", folder_name, "disk")

	# raid 1
	cmd_and_capture(expect, "storcli /c0 /vall show |grep 'Virtual Drives' -A8", folder_name, "raid1")

	# raid 2
	cmd_and_capture(expect, "storcli /c0 /eall /sall show |grep 'Drive Information' -A11", folder_name, "raid2")

	## npu 1
	cmd_and_capture(expect, "rbln-stat", folder_name, "npu1")

	## npu 2
	cmd_and_capture(expect, "rbln-stat -j | grep link", folder_name, "npu2")

	## npu 3
	cmd_and_capture(expect, "rbln-stat -L", folder_name, "npu3")




	connection.close()


## bmc 
sheet = wb["bmc"]

for row_num in range(2, sheet.max_row + 1):
  row = sheet[row_num]
  name = row[3].value
  
  if name in d:
    v = d[name]
     
    print(v)
     
    row[5].value = v["manufacturer"]
    row[6].value = v["product_name"]
    row[7].value = v["serial_number"]
	
## 엑셀 저장
wb.save('inventory.xlsx')
